"""Generación de reportes: PDF (hoja de dosificación) y Excel (kardex, producción,
rentabilidad, mermas). Toda la lógica/cálculo vive en Python; estos reportes solo
formatean datos ya calculados por los demás servicios."""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session

from app.models.produccion import OrdenProduccion
from app.models.recetas import RecetaVersion
from app.services import inventario_service, recetas_service


def hoja_dosificacion_pdf(db: Session, op_id: int) -> bytes:
    """Genera la hoja de dosificación de una OP: cantidad requerida por
    ingrediente (según receta activa y cantidad solicitada), unidad, lote
    sugerido (FEFO) y espacios para operario/hora firmados a mano."""
    op = db.query(OrdenProduccion).get(op_id)
    if op is None:
        raise ValueError("Orden de producción no encontrada")

    receta = recetas_service.receta_activa(db, op.producto_id)
    if receta is None:
        raise ValueError("El producto no tiene receta activa")

    almacen = inventario_service.obtener_o_crear_almacen(db, op.sucursal_id)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph("HOJA DE DOSIFICACIÓN", styles["Title"]),
        Paragraph(f"OP: {op.codigo}", styles["Normal"]),
        Paragraph(f"Producto: {op.producto.nombre}", styles["Normal"]),
        Paragraph(f"Tienda: {op.sucursal.nombre}", styles["Normal"]),
        Paragraph(f"Cantidad solicitada: {op.cantidad_solicitada} {op.producto.unidad}", styles["Normal"]),
        Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    data = [["Insumo", "Cant. requerida", "Unidad", "Lote sugerido (FEFO)", "Ubicación"]]
    for ingrediente in receta.ingredientes:
        cantidad_requerida = float(ingrediente.cantidad_por_unidad) * float(op.cantidad_solicitada)
        lote = inventario_service.sugerir_lote_fefo(db, ingrediente.insumo_id, almacen.id)
        data.append(
            [
                ingrediente.insumo.nombre,
                f"{cantidad_requerida:.3f}",
                ingrediente.insumo.unidad,
                lote.codigo_lote if lote else "-",
                almacen.nombre,
            ]
        )

    tabla = Table(data, repeatRows=1, colWidths=[5 * cm, 3 * cm, 2.5 * cm, 4 * cm, 3.5 * cm])
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ]
        )
    )
    elementos.append(tabla)
    elementos.append(Spacer(1, 1.5 * cm))
    elementos.append(Paragraph("Operario: ______________________   Hora: __________", styles["Normal"]))
    elementos.append(Spacer(1, 0.4 * cm))
    elementos.append(Paragraph("Supervisor: ______________________   Firma: __________", styles["Normal"]))

    doc.build(elementos)
    return buf.getvalue()


def _autofit(ws) -> None:
    for col in ws.columns:
        largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(largo + 2, 40)


def _hoja(wb: Workbook, titulo: str, encabezados: list[str], filas: list[list]):
    ws = wb.create_sheet(titulo[:31])
    ws.append(encabezados)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    for fila in filas:
        ws.append(fila)
    _autofit(ws)
    return ws


def exportar_produccion_excel(db: Session) -> bytes:
    ops = db.query(OrdenProduccion).order_by(OrdenProduccion.id.desc()).all()
    wb = Workbook()
    wb.remove(wb.active)
    _hoja(
        wb,
        "Producción",
        ["Código OP", "Tienda", "Producto", "Estado", "Cant. solicitada", "Cant. producida", "Costo real total"],
        [
            [
                op.codigo,
                op.sucursal.nombre,
                op.producto.nombre,
                op.estado.value if hasattr(op.estado, "value") else op.estado,
                float(op.cantidad_solicitada),
                float(op.cantidad_producida) if op.cantidad_producida else None,
                float(op.costo_real_total) if op.costo_real_total else None,
            ]
            for op in ops
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_kardex_excel(db: Session) -> bytes:
    from app.models.inventario import MovimientoInventario

    movs = db.query(MovimientoInventario).order_by(MovimientoInventario.id.desc()).all()
    wb = Workbook()
    wb.remove(wb.active)
    _hoja(
        wb,
        "Kardex",
        ["Fecha", "Producto", "Almacén", "Tipo", "Cantidad", "Costo unitario", "Costo total", "Referencia"],
        [
            [
                m.creado_en.strftime("%Y-%m-%d %H:%M") if m.creado_en else "",
                m.producto.nombre,
                m.almacen.nombre,
                m.tipo.value if hasattr(m.tipo, "value") else m.tipo,
                float(m.cantidad),
                float(m.costo_unitario),
                float(m.costo_total),
                m.referencia,
            ]
            for m in movs
        ],
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
